"""Extract country-level total enrollment time series from the IIE Open Doors
'All Places of Origin, Selected Years' workbook."""
import openpyxl
import pandas as pd

SHEET_NAME = "Intl Students Place of Origin"

# Countries in the shock catalog (affected) + unaffected "control" countries
# for later causal comparison.
TARGET_COUNTRIES = [
    "China", "India", "Nepal", "Afghanistan", "Iran", "Iraq", "Israel",
    "Russia", "Ukraine", "Mexico",
    "South Korea", "Vietnam", "Brazil", "Nigeria", "Canada", "Japan",
]


def load_enrollment(source_path: str) -> pd.DataFrame:
    """Parse the Open Doors workbook into a long-format country-year enrollment table."""
    wb = openpyxl.load_workbook(source_path, data_only=True)
    ws = wb[SHEET_NAME]

    header_row = [cell.value for cell in ws[3]]
    year_cols = [
        (idx, val) for idx, val in enumerate(header_row, start=1)
        if isinstance(val, str) and "/" in val
    ]

    records = []
    matched = set()
    for row in ws.iter_rows(min_row=5, max_row=294):
        name = row[1].value
        if name is None:
            continue
        name_clean = str(name).strip()
        if name_clean in TARGET_COUNTRIES:
            matched.add(name_clean)
            for col_idx, year_label in year_cols:
                val = row[col_idx - 1].value
                if val in (None, "-", ""):
                    continue
                records.append({
                    "country": name_clean,
                    "academic_year": year_label,
                    "total_enrollment": val,
                })

    missing = set(TARGET_COUNTRIES) - matched
    if missing:
        print(f"WARNING: target countries not found in sheet: {missing}")

    return pd.DataFrame(records)


if __name__ == "__main__":
    df = load_enrollment("/Users/aryan/Documents/geopolitical-student-mobility/data/raw/Census_All-Places-of-Origin-Selected-Years_OD25_Website.xlsx")
    df.to_csv("data/processed/enrollment_by_country.csv", index=False)
    print(df.shape)
    print(df.head())