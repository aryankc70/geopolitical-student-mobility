"""Extract F-1 student visa issuances by country and fiscal year from the
State Department's consolidated NIV Detail Table workbook (FY1997-2024)."""
import openpyxl
import pandas as pd
import re

# Canonical country name -> exact string(s) used in THIS source file.
# Different sources (State Dept vs IIE Open Doors) use different naming
# conventions - this harmonizes them to the canonical names used across
# the whole project's panel dataset.
COUNTRY_ALIASES = {
    "China": ["China - mainland", "China"],
    "India": ["India"],
    "Nepal": ["Nepal"],
    "Afghanistan": ["Afghanistan"],
    "Iran": ["Iran"],
    "Iraq": ["Iraq"],
    "Israel": ["Israel"],
    "Russia": ["Russia"],
    "Ukraine": ["Ukraine"],
    "Mexico": ["Mexico"],
    "South Korea": ["Korea, South"],
    "Vietnam": ["Vietnam"],
    "Brazil": ["Brazil"],
    "Nigeria": ["Nigeria"],
    "Canada": ["Canada"],
    "Japan": ["Japan"],
}
ALIAS_TO_CANONICAL = {alias: canon for canon, aliases in COUNTRY_ALIASES.items() for alias in aliases}


def find_f1_column(header_row):
    """F-1 label format varies by year: 'F-1', 'F1', etc."""
    for idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        cleaned = str(val).strip().upper().replace("-", "").replace(" ", "")
        if cleaned == "F1":
            return idx
    return None


def load_visa_data(source_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(source_path, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        f1_col = find_f1_column(header_row)

        fy_label = ws.cell(row=1, column=1).value
        year_match = re.search(r"(\d{4})", str(fy_label)) if fy_label else None
        fiscal_year = int(year_match.group(1)) if year_match else None

        if f1_col is None or fiscal_year is None:
            print(f"WARNING: skipped sheet {sheet_name} - no F1 column or year found")
            continue

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            name = row[0].value
            if name is None:
                continue
            canonical = ALIAS_TO_CANONICAL.get(str(name).strip())
            if canonical:
                val = row[f1_col - 1].value
                if val in (None, "-", ""):
                    continue
                records.append({
                    "country": canonical,
                    "fiscal_year": fiscal_year,
                    "f1_visas_issued": val,
                })

    return pd.DataFrame(records)


if __name__ == "__main__":
    df = load_visa_data("/Users/aryan/Documents/geopolitical-student-mobility/data/raw/FYs97-24_NIVDetailTable.xlsx")
    df.to_csv("data/processed/visa_by_country.csv", index=False)
    print(df.shape)