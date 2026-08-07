"""Extract country-level enrollment by academic level from the IIE Open Doors
'Places of Origin and Academic Level' workbook.

Note: the source spreadsheet's category labels change partway through the
time series - years before ~2006/07 use 'Under-graduate', 'Graduate', 'Other';
years from ~2006/07 onward split 'Other' into 'Non-Degree' and 'OPT'
(Optional Practical Training). This script recognizes all five labels and
combines Other + Non-Degree + OPT into a single 'enrollment_other' column
so pre- and post-2006/07 years stay comparable.
"""
import openpyxl
import pandas as pd

SHEET_NAME = "Places of Origin&Academic Level"

TARGET_COUNTRIES = [
    "China", "India", "Nepal", "Afghanistan", "Iran", "Iraq", "Israel",
    "Russia", "Ukraine", "Mexico",
    "South Korea", "Vietnam", "Brazil", "Nigeria", "Canada", "Japan",
]

# maps every label seen in the source to one of our three canonical buckets
LEVEL_LABEL_MAP = {
    "Under-graduate": "enrollment_undergrad",
    "Graduate": "enrollment_grad",
    "Other": "enrollment_other",
    "Non-Degree": "enrollment_other",
    "OPT": "enrollment_other",
}


def load_enrollment_by_level(source_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(source_path, data_only=True)
    ws = wb[SHEET_NAME]

    year_row = [c.value for c in ws[2]]
    level_row = [c.value for c in ws[3]]

    col_map = []
    current_year = None
    for idx, val in enumerate(year_row, start=1):
        if val is not None:
            current_year = val
        lvl = level_row[idx - 1]
        if lvl in LEVEL_LABEL_MAP:
            col_map.append((idx, current_year, LEVEL_LABEL_MAP[lvl]))

    records = []
    matched = set()
    for row in ws.iter_rows(min_row=5, max_row=303):
        name = row[0].value
        if name is None:
            continue
        name_clean = str(name).strip()
        if name_clean in TARGET_COUNTRIES:
            matched.add(name_clean)
            for col_idx, year_label, canonical_level in col_map:
                val = row[col_idx - 1].value
                if val in (None, "-", ""):
                    continue
                records.append({
                    "country": name_clean,
                    "academic_year": year_label,
                    "academic_level": canonical_level,
                    "enrollment": val,
                })

    missing = set(TARGET_COUNTRIES) - matched
    if missing:
        print(f"WARNING: not found: {missing}")

    df = pd.DataFrame(records)
    # sum Other+Non-Degree+OPT together per country-year (they were split
    # into up to 3 separate rows above, all mapped to the same canonical bucket)
    df = df.groupby(["country", "academic_year", "academic_level"], as_index=False)["enrollment"].sum()
    return df


if __name__ == "__main__":
    df = load_enrollment_by_level("data/raw/Census_Places-of-Origin-Academic-Level_OD25_Website.xlsx")
    df.to_csv("data/processed/enrollment_by_country_level.csv", index=False)
    print(df.shape)